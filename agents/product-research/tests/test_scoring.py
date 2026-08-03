from io import BytesIO

from amazon_product_research.models import (
    Candidate,
    Decision,
    ResearchConstraints,
    ResearchMode,
    ResearchResult,
    ResearchRun,
)
from amazon_product_research.report import xlsx_bytes
from amazon_product_research.scoring import decide, score_candidate
from openpyxl import load_workbook


def test_score_is_deterministic_and_explainable() -> None:
    candidate = Candidate(
        title="Storage organizer",
        price_usd=39.99,
        cost_usd=8,
        monthly_search_volume=80_000,
        review_count=120,
        trend_pct=12,
        supplier_count=15,
    )
    scored = score_candidate(candidate, ResearchConstraints())
    assert scored.score is not None
    assert 0 <= scored.score.overall <= 100
    assert scored.estimated_margin_pct is not None
    assert set(scored.score.rationale) == {
        "demand",
        "competition",
        "profitability",
        "trend",
        "differentiation",
        "supply",
        "risk",
    }


def test_missing_cost_never_produces_unconditional_go() -> None:
    candidate = score_candidate(
        Candidate(
            title="Unknown cost product",
            price_usd=39.99,
            monthly_search_volume=100_000,
            review_count=10,
            trend_pct=20,
            supplier_count=20,
        ),
        ResearchConstraints(),
    )
    assert decide([candidate], ResearchConstraints(), []) is Decision.CONDITIONAL_GO


def test_margin_hard_gate_produces_no_go() -> None:
    candidate = score_candidate(
        Candidate(
            title="Unprofitable product",
            price_usd=20,
            cost_usd=18,
            monthly_search_volume=100_000,
            review_count=10,
            trend_pct=20,
            supplier_count=20,
        ),
        ResearchConstraints(),
    )
    assert decide([candidate], ResearchConstraints(), []) is Decision.NO_GO


def test_excel_contains_full_business_workbook() -> None:
    candidate = score_candidate(
        Candidate(title="Organizer", price_usd=30, cost_usd=6), ResearchConstraints()
    )
    run = ResearchRun(
        source_text="organizer",
        mode=ResearchMode.VALIDATE,
        result=ResearchResult(candidates=[candidate], constraints=ResearchConstraints()),
    )
    workbook = load_workbook(BytesIO(xlsx_bytes(run)), read_only=True)
    required = {
        "数据来源说明",
        "执行摘要",
        "市场与趋势",
        "关键词",
        "竞品格局",
        "评论痛点",
        "单位经济模型",
        "供应商与MOQ",
        "风险",
        "评分与决策",
        "数据缺口",
    }
    assert required.issubset(workbook.sheetnames)
